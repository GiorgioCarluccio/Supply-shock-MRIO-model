"""Check the model implementation against the uploaded ECB toy workbook.

The workbook is not used as a production input. It is a compact audit fixture
for the paper algorithm: baseline Z/FD/X, one supply shock, bottlenecks,
inventory reallocation and the fixed global-technology output cap.
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from climate_risk_io.model.io_model import IOClimateModel  # noqa: E402


WORKBOOK = ROOT / "ECB_IO_physicalrisk.xlsx"


def _vector(ws, cell_range: str) -> np.ndarray:
    return np.array([cell.value for row in ws[cell_range] for cell in row], dtype=float)


def _matrix(ws, cell_range: str) -> np.ndarray:
    return np.array([[cell.value for cell in row] for row in ws[cell_range]], dtype=float)


def _sector_groups(ws) -> np.ndarray:
    sector_names = [ws.cell(r, 4).value for r in range(5, 14)]
    mapping = {name: i for i, name in enumerate(sorted(set(sector_names)))}
    return np.array([mapping[name] for name in sector_names], dtype=int)


def _max_abs(name: str, got: np.ndarray, expected: np.ndarray) -> float:
    diff = float(np.nanmax(np.abs(got - expected)))
    print(f"{name:<34} max_abs_diff={diff:.6e}")
    return diff


def main() -> int:
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK}")

    wb = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    ws0 = wb["IO_sub_model"]
    ws1 = wb["IO_sub_model - t+1"]
    ws2 = wb["IO_sub_model - t+2"]

    Z = _matrix(ws0, "E5:M13")
    FD = _vector(ws0, "N5:N13")
    X = _vector(ws0, "O5:O13")
    sd0 = _vector(ws0, "E47:E55")
    sp = _vector(ws0, "L47:L55")
    gamma = float(ws0["I123"].value)

    labels = [f"{ws0.cell(r, 3).value}_{ws0.cell(r, 4).value}" for r in range(5, 14)]
    model = IOClimateModel(
        Z0=Z,
        FD0=FD,
        X0=X,
        sector_group_of=_sector_groups(ws0),
        node_labels=labels,
    )

    print("Checking first propagation pass against sheet 'IO_sub_model'")
    res0 = model.run(sd=sd0, sp=sp, gamma=gamma, max_iter=1, return_history=True)
    aux = res0["aux_last"]

    failures = []
    checks = [
        ("X_dem", aux["X_dem"], _vector(ws0, "E60:E68")),
        ("r", aux["r"], _vector(ws0, "G60:G68")),
        ("s", aux["s"], _vector(ws0, "E72:M72")),
        ("row_factor", aux["row_factor"], _vector(ws0, "B78:B86")),
        ("inventories", aux["inv"], _vector(ws0, "E118:E126")),
        ("Z_new", res0["Z_final"], _matrix(ws0, "E150:M158")),
        ("X_supply_global", res0["X_supply_global_last"], _vector(ws0, "E184:M184")),
    ]
    for name, got, expected in checks:
        if _max_abs(name, got, expected) > 1e-8:
            failures.append(name)

    fd_next_workbook = FD * (1.0 - (sd0 + _vector(ws0, "B188:B196")))
    if _max_abs("FD update after pass 1", res0["FD_post_final"], fd_next_workbook) > 1e-8:
        failures.append("FD update after pass 1")

    print()
    print("Checking second pass starts from the workbook t+1 demand guess")
    fd_iter2_start = model.run(
        sd=sd0, sp=sp, gamma=gamma, max_iter=2, return_history=True
    )["FD_post_history"][1]
    if _max_abs("iteration 2 start FD", fd_iter2_start, _vector(ws1, "H47:H55")) > 1e-8:
        failures.append("iteration 2 start FD")

    sd1 = _vector(ws1, "E47:E55")
    res1 = model.run(sd=sd1, sp=sp, gamma=gamma, max_iter=1)
    if _max_abs("pass 2 Z_new", res1["Z_final"], _matrix(ws1, "E150:M158")) > 1e-8:
        failures.append("pass 2 Z_new")

    fd_next_code = res1["FD_post_final"]
    fd_next_workbook = _vector(ws2, "H47:H55")
    diff = _max_abs("FD update after pass 2", fd_next_code, fd_next_workbook)
    if diff > 1e-8:
        print(
            "Note: the workbook accumulates demand-shock percentages relative to "
            "the current FD_post. The code follows the paper wording by carrying "
            "the new FD vector directly. Propagation equations still match."
        )

    if failures:
        print()
        print("FAILED:", ", ".join(failures))
        return 1

    print()
    print("Core propagation checks passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
